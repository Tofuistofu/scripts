#!\usr\bin\env python
''' BatchEmail.py - Sends customized emails to clients
Reads a template from a .txt file and replaces keyword fields in the template using client information
in an Excel spreadsheet. Sends a customized email to each client. '''

import openpyxl
import smtplib
import sys

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def send_email(send_from, login_pass, send_to, cc, subject, body, attachments=None, server='smtp.gmail.com', port=587):
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = send_from
    msg['To'] = send_to 
    
    # Set up cc recipients depending on what is entered
    if isinstance(cc, list):
        recipients = [send_to] + cc
        msg['Cc'] = ','.join(cc)
    elif not cc:
        recipients = send_to
    else:
        recipients = [send_to, cc]
        msg['Cc'] = cc

    msg.attach(MIMEText(body))
    
    # Attach file attachments to email
    for files in attachments or []:
        with open(files, 'rb') as f:
            part = MIMEBase('doc', 'octet-stream')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=files)
        msg.attach(part)           

    print 'Sending mail to {}:'.format(recipients)
    
    smtpObj = smtplib.SMTP(server, port)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(send_from, login_pass)
    sendStatus = smtpObj.sendmail(send_from, recipients, msg.as_string())
    if sendStatus != {}:
        print 'An issue occured while sending email to {0}: {1}'.format(recipients, sendStatus)
    smtpObj.quit()

def main():
    try:
        sender = sys.argv[1]
        password = sys.argv[2]
    except IndexError:
        print 'Usage: -py -username -password'
        
    with open('BatchEmails_Template.txt') as f:
        template = f.read()

    wb = openpyxl.load_workbook('BatchEmails_Client Info.xlsx')
    sheet = wb.get_sheet_by_name('Client Information')
    # Read spreadsheet data
    sheet_data = []    
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        row_values = ()
        for cell in row:
            row_values += cell.value,
        sheet_data.append(row_values)
    # Send customized email for each row of client data in spreadsheet
    clients = 0
    for client_info in sheet_data[1:]:
        client_email = client_info[2]
        rep_email = client_info[4]

        email_contents = template
        wordreplace = zip(sheet_data[0], client_info)
        for old, new in wordreplace:
            email_contents = email_contents.replace(old, str(new))
        subject, body = email_contents.split('\n', 1)
        # Adding attachments for each individual client will go here
        send_email(sender, password, client_email, rep_email, subject, body, attachment) 
        clients += 1

    print 'Task Completed. \n{} emails have been sent to clients.'.format(clients)

    
if __name__ == '__main__':
    main()
